#!/usr/bin/env python3
"""
优化版 Excel 导出脚本 - 全面美化表格格式
功能：
1. 列宽自动调整
2. 添加样式（颜色、边框、字体）
3. 增加信息列（Jira 链接、最后评论时间、场景类型）
4. 优化列顺序
5. 表头冻结
6. 条件格式（停滞天数标红）
"""

import json
import os
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Fill, PatternFill, Border, Side, Alignment, Color
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils.dataframe import dataframe_to_rows

# 配置
REPORTS_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'data', 'reports')

# 样式定义
class Styles:
    # 字体
    HEADER_FONT = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
    TITLE_FONT = Font(name='微软雅黑', size=11, bold=True)
    NORMAL_FONT = Font(name='微软雅黑', size=10)
    LINK_FONT = Font(name='微软雅黑', size=10, color='0066CC', underline='single')
    
    # 填充色
    HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')  # 蓝色表头
    ALT_ROW_FILL = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')  # 灰色隔行
    STAGNATION_WARNING = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # 红色警告
    STAGNATION_CAUTION = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # 黄色注意
    
    # 边框
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 对齐
    CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=False)
    LEFT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)

def load_consolidated_results():
    """加载合并后的 JSON 结果"""
    json_files = sorted(
        [f for f in os.listdir(REPORTS_DIR) if f.endswith('_all_suggestions.json')],
        reverse=True
    )
    
    if not json_files:
        print("❌ 未找到合并后的 JSON 文件")
        return []
    
    latest_file = json_files[0]
    json_path = os.path.join(REPORTS_DIR, latest_file)
    
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    print(f"📊 加载文件：{latest_file}")
    
    # 处理嵌套结构
    suggestions = data.get('suggestions', []) if isinstance(data, dict) else data
    print(f"📈 工单数量：{len(suggestions)}")
    
    return suggestions

def extract_last_comment_time(raw_issue):
    """提取最后一条人类评论的时间"""
    comments = raw_issue.get('comments', [])
    
    # 过滤掉系统评论
    human_comments = [
        c for c in comments 
        if 'Automation' not in c.get('author', {}).get('displayName', '')
        and 'Jira' not in c.get('author', {}).get('displayName', '')
    ]
    
    if not human_comments:
        return None
    
    last_comment = human_comments[-1]
    created = last_comment.get('created', '')
    
    # 格式化日期
    if created:
        try:
            dt = datetime.strptime(created[:10], '%Y-%m-%d')
            return dt.strftime('%Y-%m-%d')
        except:
            return created[:10]
    
    return None

def clean_suggestion_text(text):
    """清理 LLM 输出的 artifacts，如 (1) 空格等"""
    if not text:
        return text
    
    import re
    # 移除 (1) 模式
    text = re.sub(r'\(1\)', '', text)
    # 移除多余空格（中英文混排时的空格）
    text = re.sub(r'\s+', '', text)
    # 清理首尾空格
    text = text.strip()
    
    return text

def create_workbook(suggestions):
    """创建工作簿并设置样式"""
    wb = Workbook()
    ws = wb.active
    ws.title = "工单回复建议"
    
    # 定义列（优化后的顺序）
    columns = [
        ('序号', 6),      # A
        ('工单号', 12),   # B - 重要信息靠前
        ('Jira 链接', 18), # C - 新增超链接
        ('负责人', 12),   # D
        ('组织', 15),     # E - 新增组织信息
        ('停滞天数', 10), # F - 重要指标
        ('最后评论', 12), # G - 新增时间
        ('场景类型', 15), # H - 新增 AI 判断
        ('建议评论', 60), # I - 核心内容，加宽
        ('是否发表', 10), # J - Y/N 确认
        ('备注', 20),     # K - 修改意见
    ]
    
    # 设置表头
    for col_idx, (header, _) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Styles.HEADER_FONT
        cell.fill = Styles.HEADER_FILL
        cell.border = Styles.THIN_BORDER
        cell.alignment = Styles.CENTER_ALIGN
    
    # 设置列宽
    for col_idx, (_, width) in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # 填充数据
    for row_idx, suggestion in enumerate(suggestions, 2):
        issue_key = suggestion['issue_key']
        raw_issue = suggestion.get('raw_issue', {})
        
        # 提取额外信息
        last_comment_time = extract_last_comment_time(raw_issue)
        organization = raw_issue.get('fields', {}).get('customfield_10002', [])
        org_name = organization[0].get('name', '未知') if organization else '未知'
        
        # Jira 链接（超链接格式）
        jira_link = f"https://streamaxamerica.atlassian.net/browse/{issue_key}"
        
        # 清理建议文本（移除 (1) 等 artifacts）
        raw_suggestion = suggestion.get('suggestion', '')
        clean_suggestion = clean_suggestion_text(raw_suggestion)
        
        # 数据行
        row_data = [
            row_idx - 1,  # 序号
            issue_key,
            jira_link,  # 超链接
            suggestion.get('assignee', 'Unknown'),
            org_name,
            suggestion.get('stagnation_days', 0),
            last_comment_time or '-',
            suggestion.get('scene_type', '未知'),
            clean_suggestion,  # 清理后的建议
            '',  # Y/N 列（留空待用户填写）
            '',  # 备注列（留空待用户填写）
        ]
        
        # 写入数据
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Styles.NORMAL_FONT
            cell.border = Styles.THIN_BORDER
            
            # 特殊列样式
            if col_idx == 3:  # Jira 链接
                cell.font = Styles.LINK_FONT
                cell.alignment = Styles.CENTER_ALIGN
            elif col_idx == 9:  # 建议评论
                cell.alignment = Styles.LEFT_ALIGN
            elif col_idx in [10, 11]:  # Y/N 和备注
                cell.alignment = Styles.CENTER_ALIGN
            else:
                cell.alignment = Styles.CENTER_ALIGN
        
        # 隔行变色
        if row_idx % 2 == 0:
            for col_idx in range(1, len(columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = Styles.ALT_ROW_FILL
        
        # 条件格式：停滞天数标红
        stagnation_days = suggestion.get('stagnation_days', 0)
        if stagnation_days >= 21:
            for col_idx in range(1, len(columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = Styles.STAGNATION_WARNING
        elif stagnation_days >= 14:
            for col_idx in range(1, len(columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = Styles.STAGNATION_CAUTION
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    # 添加标题行
    ws.insert_rows(1)
    title_cell = ws.cell(row=1, column=1, value=f"LLM 工单回复建议 - {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    title_cell.font = Font(name='微软雅黑', size=16, bold=True, color='1F4E79')
    title_cell.alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:K1')
    
    # 添加说明行
    ws.insert_rows(2)
    info_text = "使用说明：1. 在'是否发表'列填写 Y/N  2. 可在'备注'列添加修改意见  3. 红色=停滞≥21 天，黄色=停滞≥14 天"
    info_cell = ws.cell(row=2, column=1, value=info_text)
    info_cell.font = Font(name='微软雅黑', size=9, italic=True, color='666666')
    info_cell.alignment = Alignment(horizontal='left')
    ws.merge_cells('A2:K2')
    ws.row_dimensions[2].height = 25
    
    # 调整行高
    ws.row_dimensions[1].height = 30  # 标题行
    ws.row_dimensions[3].height = 25  # 表头
    
    return wb

def main():
    """主函数"""
    print("🚀 开始生成优化版 Excel 报告...")
    
    # 加载数据
    suggestions = load_consolidated_results()
    
    if not suggestions:
        print("❌ 没有数据可导出")
        return
    
    # 创建工作簿
    wb = create_workbook(suggestions)
    
    # 保存文件
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"{timestamp}_LLM 工单回复建议_优化版.xlsx"
    output_path = os.path.join(REPORTS_DIR, filename)
    
    wb.save(output_path)
    
    print(f"✅ Excel 已生成：{filename}")
    print(f"📂 路径：{output_path}")
    print(f"📊 工单数量：{len(suggestions)}")
    
    # 同时推送到 Windows
    print("\n📤 正在推送到 Windows 共享...")
    import subprocess
    
    # 复制到临时路径（避免中文路径问题）
    temp_path = f"/tmp/LLM_工单回复建议_{timestamp}.xlsx"
    subprocess.run(['cp', output_path, temp_path], check=True)
    
    # 推送
    cmd = f'''smbclient "//192.168.60.20/workshop" -U "Emrys Liang" << 'EOF'
cd jiraAss
lcd /tmp
prompt OFF
put {os.path.basename(temp_path)} {os.path.basename(output_path)}
ls {os.path.basename(output_path)}
exit
EOF
'''
    
    result = subprocess.run(cmd, shell=True, capture_output=True, text=True)
    
    if os.path.basename(output_path) in result.stdout:
        print(f"✅ 推送成功：E:\\WSL\\workshop\\jiraAss\\{os.path.basename(output_path)}")
    else:
        print(f"⚠️ 推送可能失败，请手动检查")
        print(result.stderr)

if __name__ == '__main__':
    main()
