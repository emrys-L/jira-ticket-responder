#!/usr/bin/env python3
"""
小批量测试：分析前 10 个活跃工单并生成 Excel
"""

import json
import sys
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).parent.parent))

from scripts.llm_analyzer import analyze_issue, load_all_issues

def main():
    print("="*100)
    print("📊 Jira Ticket Responder - 小批量测试（前 10 个工单）")
    print("="*100)
    print()
    
    # 加载所有工单
    print("📋 加载工单列表...")
    all_issues = load_all_issues()
    print(f"   总工单数：{len(all_issues)}")
    print(f"   本次测试：前 10 个")
    print()
    
    # 只分析前 10 个
    test_issues = all_issues[:10]
    
    # 批量分析
    results = []
    
    for i, issue_key in enumerate(test_issues, 1):
        print(f"[{i}/{len(test_issues)}] 分析 {issue_key}...", end=" ")
        
        result = analyze_issue(issue_key)
        
        if result and result.get('issue_key'):
            print("✅ 成功")
            results.append(result)
        else:
            print("⏭️  跳过/失败")
    
    print()
    print("="*100)
    print(f"📊 结果：成功 {len(results)}/{len(test_issues)}")
    print("="*100)
    print()
    
    if not results:
        print("⚠️  没有生成任何建议，退出")
        return
    
    # 生成 Excel
    print("📝 生成 Excel...")
    wb = Workbook()
    ws = wb.active
    ws.title = "工单回复建议"
    
    # 表头
    headers = ['工单号', '负责人', '停滞天数', '建议评论', '是否发表 (Y/N)', '备注']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # 填充数据
    for row, result in enumerate(results, 2):
        # 提取负责人
        suggestion_text = result.get('suggestion', '')
        assignee = 'N/A'
        if suggestion_text.startswith('@'):
            parts = suggestion_text.split()
            if parts:
                assignee = parts[0][1:]
        
        ws.cell(row=row, column=1, value=result.get('issue_key'))
        ws.cell(row=row, column=2, value=assignee)
        ws.cell(row=row, column=3, value=result.get('stagnation_days', 0))
        ws.cell(row=row, column=4, value=suggestion_text)
        ws.cell(row=row, column=5, value='')  # Y/N
        ws.cell(row=row, column=6, value='')  # 备注
    
    # 列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 80
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 40
    
    # 保存
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = f'data/reports/{timestamp}_batch10_suggestions.xlsx'
    wb.save(output_file)
    
    print(f"✅ Excel 已保存：{output_file}")
    print()
    print("📋 使用说明:")
    print("   1. 打开 Excel 文件")
    print("   2. 在 E 列填写 Y(发表) 或 N(不发表)")
    print("   3. 可在 F 列添加修改意见")
    print()
    
    # JSON 结果
    json_output = f'data/reports/{timestamp}_batch10_suggestions.json'
    with open(json_output, 'w', encoding='utf-8') as f:
        json.dump({
            'timestamp': datetime.now().isoformat(),
            'total': len(test_issues),
            'success': len(results),
            'suggestions': results
        }, f, ensure_ascii=False, indent=2)
    
    print(f"✅ JSON 已保存：{json_output}")
    print()
    
    # 打印建议预览
    print("="*100)
    print("💡 建议预览")
    print("="*100)
    for i, r in enumerate(results, 1):
        print(f"\n{i}. {r['issue_key']} ({r.get('stagnation_days', 0)}天)")
        print(f"   {r['suggestion'][:100]}...")

if __name__ == '__main__':
    main()
