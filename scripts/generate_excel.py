#!/usr/bin/env python3
"""
批量分析所有工单并生成 Excel（带 Y/N 确认列）
"""

import json
import sys
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook

# 添加项目根目录到路径
sys.path.insert(0, str(Path(__file__).parent.parent))

from scripts.llm_analyzer import analyze_issue, load_all_issues

def main():
    print("="*100)
    print("📊 Jira Ticket Responder - 批量分析 + Excel 导出")
    print("="*100)
    print()
    
    # 加载所有工单
    print("📋 加载工单列表...")
    all_issues = load_all_issues()
    print(f"   总工单数：{len(all_issues)}")
    print()
    
    # 批量分析
    results = []
    failed = []
    skipped = []
    
    for i, issue_key in enumerate(all_issues, 1):
        print(f"[{i}/{len(all_issues)}] 分析 {issue_key}...", end=" ")
        
        result = analyze_issue(issue_key)
        
        if result is None:
            print("⏭️  已排除")
            skipped.append(issue_key)
        elif result.get('skipped'):
            print("✅ 已处理（跳过 LLM）")
            results.append(result)
        elif result.get('issue_key'):
            print("✅ 成功")
            results.append(result)
        else:
            print("❌ 失败")
            failed.append(issue_key)
    
    print()
    print("="*100)
    print("📊 分析统计")
    print("="*100)
    print(f"总工单数：{len(all_issues)}")
    print(f"成功：{len(results)}")
    print(f"失败：{len(failed)}")
    print(f"已排除：{len(skipped)}")
    print()
    
    # 生成 Excel
    if not results:
        print("⚠️  没有生成任何建议，退出")
        return
    
    print("📝 生成 Excel...")
    wb = Workbook()
    ws = wb.active
    ws.title = "工单回复建议"
    
    # 表头
    headers = [
        '工单号',
        '负责人',
        '停滞天数',
        '建议评论',
        '是否发表 (Y/N)',
        '备注',
        '场景类型',
        'Prompt'
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # 填充数据
    for row, result in enumerate(results, 2):
        ws.cell(row=row, column=1, value=result.get('issue_key'))
        ws.cell(row=row, column=2, value=result.get('assignee', 'N/A'))
        ws.cell(row=row, column=3, value=result.get('stagnation_days', 0))
        ws.cell(row=row, column=4, value=result.get('suggestion', ''))
        ws.cell(row=row, column=5, value='')  # Y/N 列，留空给用户填写
        ws.cell(row=row, column=6, value='')  # 备注列，留空
        ws.cell(row=row, column=7, value=result.get('scenario', ''))
        ws.cell(row=row, column=8, value=result.get('prompt', '')[:32000])  # Excel 最大 32767 字符
    
    # 调整列宽
    ws.column_dimensions['A'].width = 15   # 工单号
    ws.column_dimensions['B'].width = 15   # 负责人
    ws.column_dimensions['C'].width = 10   # 停滞天数
    ws.column_dimensions['D'].width = 80   # 建议评论
    ws.column_dimensions['E'].width = 15   # Y/N
    ws.column_dimensions['F'].width = 30   # 备注
    ws.column_dimensions['G'].width = 15   # 场景
    ws.column_dimensions['H'].width = 50   # Prompt
    
    # 保存文件
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = f'data/reports/{timestamp}_reply_suggestions.xlsx'
    wb.save(output_file)
    
    print(f"✅ Excel 已保存：{output_file}")
    print()
    print("📋 使用说明:")
    print("   1. 打开 Excel 文件")
    print("   2. 在 E 列（是否发表）填写 Y 或 N")
    print("   3. 可在 F 列（备注）添加修改意见")
    print("   4. 保存后，运行以下命令批量发表:")
    print("      python3 scripts/publish_comments.py --excel {output_file}")
    print()
    
    # 同时保存 JSON 结果
    json_file = f'data/reports/{timestamp}_reply_suggestions.json'
    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump({
            'timestamp': datetime.now().isoformat(),
            'total': len(all_issues),
            'success': len(results),
            'failed': len(failed),
            'skipped': len(skipped),
            'suggestions': results
        }, f, ensure_ascii=False, indent=2)
    
    print(f"✅ JSON 已保存：{json_file}")

if __name__ == '__main__':
    main()
